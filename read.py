import openpyxl as op
import os

filename = 'lanet.xlsx'
subcategories_dict = {}

wb = op.load_workbook(filename, data_only=True)
sheet = wb.active

max_rows = sheet.max_row

print(max_rows)

for i in range(2, max_rows+1):
    sku = sheet.cell(row=i, column=1).value
    subcategory = sheet.cell(row=i, column=2). value
    if not sku:
        continue

    if subcategory not in subcategories_dict:
        subcategories_dict[subcategory] = [sku]
    else:
        subcategories_dict[subcategory].append(sku)

sorteddict = dict(sorted(subcategories_dict.items()))

# Создаю новую книгу
book = op.Workbook()

# Удаляю таблицу по умолчанию
book.remove(book.active)

# Создаю новую таблицу
sheet = book.create_sheet("Subcategories")

# Записываю данные в Excel
sheet.append(["Subcategory", "SKUs"])  # Заголовки
for key, value in sorteddict.items():
    string_values = ', '.join(value)
    sheet.append([key, string_values])  # Добавляю данные построчно

# Сохраняю Excel-файл
file_path = os.path.join(os.getcwd(), "final.xlsx")
book.save(file_path)
print(f"Файл успешно создан: {file_path}")

# Сохраняю данные в текстовый файл
with open('subcategories.txt', 'w') as myfile:
    for key, value in sorteddict.items():
        string_values = ', '.join(value)
        string_to_write = key + ' = ' + string_values + '\n'
        myfile.write(string_to_write)
print("Файл subcategories.txt успешно создан.")