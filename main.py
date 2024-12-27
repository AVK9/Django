import openpyxl
from examples.faker_data import data_samples
import os

def example():
    """
    Создание xlsx файла и запись в него
    """
    # создаю книгу
    book = openpyxl.Workbook()

    # по умолчанию создается с таблицей Sheet
    book.remove(book.active)

    # создаю таблицы
    sheet_1 = book.create_sheet("Коллеги")
    sheet_2 = book.create_sheet("Клиенты")
    sheet_3 = book.create_sheet("Черный список", 0)  # таблица будет первой

    for sheet in book.worksheets:  # перебираю таблицы
        for row in data_samples():  # получаю данные
            sheet.append(row)  # записываю данные в строки таблиц

    # Сохранение файла
    file_path = os.path.join(os.getcwd(), "test.xlsx")
    book.save(file_path)

    # Проверка существования файла
    if os.path.exists(file_path):
        print(f"Файл успешно создан и сохранён: {file_path}")
    else:
        print(f"Ошибка: файл не был сохранён в {file_path}")

if __name__ == "__main__":
    example()
